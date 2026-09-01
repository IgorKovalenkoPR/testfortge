"""A 2.5 MB upload cost 365 MB and two minutes, on a 512 MB dyno.

``engine/imports.py`` is the one large parser in this product fed directly by
a browser upload, and it was the one with no bounds at all. Measured on a
file of 200 000 rows that ``MAX_CONTENT_LENGTH`` (64 MB) would have accepted
twenty-five times over:

    load_workbook + list(iter_rows())     175 MB peak    56 s
    read_only=True, streamed               16 MB         47 s
    read_only=True + a 20 000-row cap       2 MB        5.8 s
    read_headers — for *one row*          190 MB         75 s

The last line is the one that turns this from a hypothetical into a defect:
``read_headers`` exists to fill the column-mapping form (E4.8), wants the
header row, and paid a full parse for it. The upload route calls it and then
parses again — so one 2.5 MB file cost roughly 365 MB of peak allocation and
over two minutes, on a single-worker dyno behind a proxy that closes idle
connections at 30 seconds. The worker dies and takes every other request on
it with it, which is a signed-in member's ordinary upload doing it.

Refuted while measuring, and pinned below so nobody re-measures it: the
sheet-``<dimension>`` bomb does not work here. A tiny file declaring
``A1:XFD1048576`` costs nothing in either mode, and read-only yields only the
rows that exist — which is what lets the cap count yielded rows without
worrying about padding.

The third instance of this family in the codebase, and the first one nobody
had guarded: ``engine/allure_ingest.py`` caps member count and member size
and explains why, and ``engine/walkthrough_stats.py`` gained a read cap
earlier in this series.
"""
from __future__ import annotations

import csv
import zipfile
from unittest import mock

import pytest

from engine import imports

pytestmark = pytest.mark.usefixtures()


def _write_xlsx(path, rows):
    """A real workbook, via openpyxl, so the reader is exercised for real."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


def _write_xlsx_with_declared_dimension(path, dimension):
    """A hand-rolled two-row sheet that *claims* to be enormous.

    openpyxl will not write a dishonest ``<dimension>``, and the dishonest
    one is the whole point of the test.
    """
    head = ('<row r="1"><c r="A1" t="inlineStr"><is><t>id</t></is></c>'
            '<c r="B1" t="inlineStr"><is><t>summary</t></is></c></row>')
    body = ('<row r="2"><c r="A2" t="inlineStr"><is><t>TC-1</t></is></c>'
            '<c r="B2" t="inlineStr"><is><t>Verify a thing</t></is></c></row>')
    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    pkg = "http://schemas.openxmlformats.org/package/2006/relationships"
    parts = {
        "[Content_Types].xml":
            '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats'
            '.org/package/2006/content-types"><Default Extension="rels" '
            'ContentType="application/vnd.openxmlformats-package.relationships'
            '+xml"/><Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml" ContentType="application/'
            'vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/worksheets/sheet1.xml" ContentType='
            '"application/vnd.openxmlformats-officedocument.spreadsheetml.'
            'worksheet+xml"/></Types>',
        "_rels/.rels":
            f'<?xml version="1.0"?><Relationships xmlns="{pkg}">'
            f'<Relationship Id="rId1" Type="{rel}/officeDocument" '
            f'Target="xl/workbook.xml"/></Relationships>',
        "xl/workbook.xml":
            f'<?xml version="1.0"?><workbook xmlns="{ns}" xmlns:r="{rel}">'
            f'<sheets><sheet name="S" sheetId="1" r:id="rId1"/></sheets>'
            f'</workbook>',
        "xl/_rels/workbook.xml.rels":
            f'<?xml version="1.0"?><Relationships xmlns="{pkg}">'
            f'<Relationship Id="rId1" Type="{rel}/worksheet" '
            f'Target="worksheets/sheet1.xml"/></Relationships>',
        "xl/worksheets/sheet1.xml":
            f'<?xml version="1.0"?><worksheet xmlns="{ns}">'
            f'<dimension ref="{dimension}"/><sheetData>{head}{body}'
            f'</sheetData></worksheet>',
    }
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, body_xml in parts.items():
            archive.writestr(name, body_xml)
    return path


def _rows(count):
    return ([["id", "summary"]]
            + [[f"TC-{i}", f"Verify row {i}"] for i in range(1, count + 1)])


@pytest.fixture
def small_cap(monkeypatch):
    """A cap of 5, so the tests stay fast and the arithmetic stays visible."""
    monkeypatch.setattr(imports, "MAX_IMPORT_ROWS", 5)
    return 5


class TestTheRowCap:

    def test_a_file_at_the_cap_is_parsed_whole(self, tmp_path, small_cap):
        """Exactly at the limit is inside it. A cap that refused the boundary
        would be a cap of four described as a cap of five."""
        path = _write_xlsx(tmp_path / "at.xlsx", _rows(small_cap))
        header, rows = imports._read_xlsx(str(path))
        assert header[:2] == ["id", "summary"]
        assert len(rows) == small_cap

    def test_one_row_past_the_cap_is_refused(self, tmp_path, small_cap):
        path = _write_xlsx(tmp_path / "over.xlsx", _rows(small_cap + 1))
        with pytest.raises(imports.ImportTooLarge):
            imports._read_xlsx(str(path))

    def test_the_refusal_says_what_to_do(self, tmp_path, small_cap):
        """The route flashes ``str(exc)`` straight at the operator, so the
        message is the whole user interface for this case."""
        path = _write_xlsx(tmp_path / "over.xlsx", _rows(small_cap + 1))
        with pytest.raises(imports.ImportTooLarge) as raised:
            imports._read_xlsx(str(path))
        message = str(raised.value)
        assert "5" in message
        assert "Split it" in message

    def test_it_refuses_rather_than_truncating(self, tmp_path, small_cap):
        """The choice this cap makes, asserted as a choice. Half a test pack
        that looks whole is worse than an upload that did not happen — the
        rows that vanished are indistinguishable from rows nobody wrote."""
        path = _write_xlsx(tmp_path / "over.xlsx", _rows(small_cap * 4))
        with pytest.raises(imports.ImportTooLarge):
            imports._read_xlsx(str(path))

    def test_the_csv_reader_is_capped_too(self, tmp_path, small_cap):
        path = tmp_path / "over.csv"
        with path.open("w", newline="", encoding="utf-8") as handle:
            csv.writer(handle).writerows(_rows(small_cap + 2))
        with pytest.raises(imports.ImportTooLarge):
            imports._read_csv(str(path))

    def test_a_small_csv_still_reads(self, tmp_path, small_cap):
        path = tmp_path / "ok.csv"
        with path.open("w", newline="", encoding="utf-8") as handle:
            csv.writer(handle).writerows(_rows(2))
        header, rows = imports._read_csv(str(path))
        assert header[:2] == ["id", "summary"]
        assert len(rows) == 2

    # There is deliberately no test that reloads the module to prove the
    # env var is read. It needs ``importlib.reload(imports)``, and a reload
    # hands out *new* class objects — so ``except ImportTooLarge`` in the
    # route would quietly stop matching the exception the reloaded reader
    # raises. The same hazard cost this project seventeen unrelated failures
    # once, under the coverage command CI gates on, from a reload in a
    # fixture. The constant's definition is one ``os.environ.get`` of the
    # shape used all over this codebase; the value it defaults to is what
    # actually needs asserting, and that is the test below.

    def test_the_default_clears_the_largest_real_pack(self):
        """4 808 cases is the measured reference plan this project's house
        style was derived from. A cap under it would refuse real work."""
        assert imports.MAX_IMPORT_ROWS >= 4808


class TestTheReaderStreams:

    def test_the_workbook_is_opened_read_only(self, tmp_path):
        """The 175 MB → 16 MB half of the fix. Asserted on the call rather
        than on the source, because a comment saying ``read_only`` is not a
        keyword argument."""
        path = _write_xlsx(tmp_path / "s.xlsx", _rows(3))
        seen = {}
        real = imports.load_workbook

        def _spy(*args, **kwargs):
            seen.update(kwargs)
            return real(*args, **kwargs)

        with mock.patch.object(imports, "load_workbook", _spy):
            imports._read_xlsx(str(path))
        assert seen.get("read_only") is True
        assert seen.get("data_only") is True

    def test_reading_the_headers_does_not_read_the_sheet(self, tmp_path):
        """The 190 MB line. ``read_headers`` went through ``_read_xlsx`` and
        threw away everything but ``[0]``; making that path explode is the
        cleanest proof it no longer takes it."""
        path = _write_xlsx(tmp_path / "h.xlsx", _rows(3))

        def _explode(*args, **kwargs):
            raise AssertionError("read_headers parsed the whole sheet again")

        with mock.patch.object(imports, "_read_xlsx", _explode):
            assert imports.read_headers(str(path), "h.xlsx") == \
                ["id", "summary"]

    def test_the_headers_of_an_oversized_file_still_read(self, tmp_path,
                                                        small_cap):
        """Which is the point of separating them: the mapping form can still
        tell the operator what columns their too-large file has."""
        path = _write_xlsx(tmp_path / "big.xlsx", _rows(small_cap * 3))
        assert imports.read_headers(str(path), "big.xlsx") == \
            ["id", "summary"]

    def test_an_inflated_dimension_does_not_trip_the_cap(self, tmp_path,
                                                        small_cap):
        """The refuted hypothesis, kept as a test.

        A tiny file declaring ``A1:XFD1048576`` was the obvious way to make
        this reader allocate a million rows. It is not: read-only mode yields
        only the rows that exist, and normal mode did not allocate per the
        declared dimension either. Pinned because the cap counts *yielded*
        rows, which is only safe while that holds.
        """
        path = _write_xlsx_with_declared_dimension(
            tmp_path / "wide.xlsx", "A1:XFD1048576")
        header, rows = imports._read_xlsx(str(path))
        assert header[:2] == ["id", "summary"]
        assert len(rows) == 1


class TestTheUploadRouteExplainsIt:

    def test_an_oversized_upload_is_a_flash_not_a_500(self, client, tmp_path,
                                                     monkeypatch):
        """``ImportTooLarge`` is a ``ValueError`` so the route's existing
        ``except Exception`` turns it into the message it already shows. The
        alternative — a new exception type nobody catches — is a 500 with the
        row count in the logs and nothing on screen."""
        monkeypatch.setattr(imports, "MAX_IMPORT_ROWS", 3)
        path = _write_xlsx(tmp_path / "over.xlsx", _rows(9))
        with path.open("rb") as handle:
            response = client.post(
                "/test-cases/upload",
                data={"upload_file": (handle, "over.xlsx"),
                      "upload_mode": "replace"},
                content_type="multipart/form-data",
                follow_redirects=False)
        assert response.status_code in (302, 303), response.status_code
        with client.session_transaction() as sess:
            flashed = " ".join(str(m) for m in sess.get("_flashes", []))
        assert "3" in flashed and "Split it" in flashed, flashed

    def test_an_ordinary_upload_still_lands(self, client, tmp_path):
        """The control. Everything above would pass on an import that refused
        every file."""
        path = _write_xlsx(tmp_path / "ok.xlsx", _rows(3))
        with path.open("rb") as handle:
            response = client.post(
                "/test-cases/upload",
                data={"upload_file": (handle, "ok.xlsx"),
                      "upload_mode": "replace"},
                content_type="multipart/form-data",
                follow_redirects=False)
        assert response.status_code in (302, 303)
        with client.session_transaction() as sess:
            flashed = " ".join(str(m) for m in sess.get("_flashes", []))
        assert "Split it" not in flashed, flashed
