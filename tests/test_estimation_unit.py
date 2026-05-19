"""Unit tests for the QA effort estimator — formulas must match the
reference template `Example of Manual QA.Estimation.v1.0.xlsx`."""
import math

from engine.qa_estimator import (
    Feature, compute_estimation, compute_features_hours,
    features_from_text, features_from_site_analysis,
    PM_OVERHEAD, BUG_REPORT_RATE, COMPATIBILITY_RATE, MAX_TESTING_STRETCH,
)


def _close(a, b, tol=1e-6):
    return abs(a - b) <= tol


# ---- Features!B40 formula ----

def test_features_hours_formula_matches_reference():
    # From reference: sample total test cases = 250 → int(1.12*250*5/60)+1 = 24
    # 1.12 * 250 * 5 / 60 = 23.333… → int = 23 → +1 = 24
    assert compute_features_hours(250) == 24

def test_features_hours_buffer_and_minutes():
    # 1.10 * 100 * 10 / 60 = 18.333… → 18 → +1 = 19
    assert compute_features_hours(100, minutes_per_tc=10, buffer=1.10) == 19


# ---- Core estimation with a known total_tc ----

def _baseline_result(total_tc: int = 250, rate: float = 30.0, extra: int = 9):
    # Simulate a project with a single feature summing to total_tc
    features = [Feature(name="Module", test_cases=total_tc)]
    return compute_estimation(
        features=features, rate_usd=rate,
        additional_platforms=extra, minutes_per_tc=5, buffer=1.12,
        project_name="Test", primary_platform="Windows 10",
    )


def test_checklist_formula():
    r = _baseline_result(250)
    # Checklist MIN = Features!B40 ; MAX = 1.5 * MIN
    chk = next(t for t in r.tasks if t.key == "checklist")
    assert chk.min_h == 24
    assert chk.max_h == 1.5 * 24

def test_functional_equals_checklist():
    r = _baseline_result(250)
    chk = next(t for t in r.tasks if t.key == "checklist")
    func = next(t for t in r.tasks if t.key == "functional")
    assert func.min_h == chk.min_h and func.max_h == chk.max_h

def test_regression_equals_functional_with_expected_g14():
    r = _baseline_result(250)
    func = next(t for t in r.tasks if t.key == "functional")
    reg = next(t for t in r.tasks if t.key == "regression")
    assert reg.min_h == func.min_h and reg.max_h == func.max_h
    # Per reference G17 = G14
    assert reg.expected_h == func.expected_h

def test_bug_report_rate():
    r = _baseline_result(250)
    func = next(t for t in r.tasks if t.key == "functional")
    reg = next(t for t in r.tasks if t.key == "regression")
    bug = next(t for t in r.tasks if t.key == "bug_report")
    assert _close(bug.min_h, (func.min_h + reg.min_h) * BUG_REPORT_RATE)
    assert _close(bug.max_h, (func.max_h + reg.max_h) * BUG_REPORT_RATE)

def test_bug_recheck_equals_bug_report():
    r = _baseline_result(250)
    bug = next(t for t in r.tasks if t.key == "bug_report")
    rec = next(t for t in r.tasks if t.key == "bug_recheck")
    assert bug.min_h == rec.min_h and bug.max_h == rec.max_h

def test_compatibility_rate():
    r = _baseline_result(250)
    chk = next(t for t in r.tasks if t.key == "checklist")
    com = next(t for t in r.tasks if t.key == "compatibility")
    assert _close(com.min_h, chk.min_h * COMPATIBILITY_RATE)
    assert _close(com.max_h, chk.max_h * COMPATIBILITY_RATE)


# ---- Totals ----

def test_one_platform_total_is_sum_plus_pm():
    r = _baseline_result(250)
    core_min = sum(t.min_h for t in r.tasks[:7])  # rows 11..17
    assert _close(r.one_plat_min, core_min)
    assert _close(r.one_plat_pm_min, core_min * PM_OVERHEAD)
    assert _close(r.one_plat_total_min, core_min * (1 + PM_OVERHEAD))

def test_full_compat_uses_additional_platforms():
    r9 = _baseline_result(250, extra=9)
    r1 = _baseline_result(250, extra=1)
    compat_min = next(t for t in r9.tasks if t.key == "compatibility").min_h
    # full - one_platform = compat_min * N
    assert _close(r9.full_min - r9.one_plat_min, compat_min * 9)
    assert _close(r1.full_min - r1.one_plat_min, compat_min * 1)


# ---- Cost ----

def test_cost_equals_total_times_rate():
    r = _baseline_result(250, rate=40.0)
    assert _close(r.cost_one_expected, r.one_plat_total_expected * 40.0)
    assert _close(r.cost_full_expected, r.full_total_expected * 40.0)


# ---- Feature parsing ----

def test_features_from_text_picks_bullets():
    txt = """Take Notes
- Writing and editing notes
- Audio transcription
Notifications
- Calendar integration
"""
    feats = features_from_text(txt)
    names = [f.name for f in feats if not f.is_section]
    assert any("Writing and editing" in n for n in names)
    assert any("Calendar integration" in n for n in names)
    assert all(f.test_cases > 0 for f in feats if not f.is_section)

def test_features_from_site_uses_pages():
    class _Page:
        def __init__(self, title, forms=0, buttons=0, nav=0):
            self.title = title; self.h1 = title; self.url = ""
            self.forms = [{} for _ in range(forms)]
            self.buttons = [str(i) for i in range(buttons)]
            self.nav_links = [str(i) for i in range(nav)]
    class _Analysis:
        pages = [_Page("Login page", forms=1, buttons=2),
                 _Page("Dashboard", forms=0, buttons=3, nav=4)]
        features_detected = []

    feats = features_from_site_analysis(_Analysis())
    assert len(feats) == 2
    assert all(f.test_cases >= 4 for f in feats)


# ---- Editable coefficients (relative / per-run rates) ----

class TestEditableCoefficients:
    """Every coefficient used by the estimation formulas must be overridable
    per run so users can express their own risk / PM profile."""

    def _run(self, **overrides):
        return compute_estimation(
            features=[Feature(name="Module", test_cases=250)],
            rate_usd=30.0,
            additional_platforms=9,
            minutes_per_tc=5,
            buffer=1.12,
            project_name="Test",
            **overrides,
        )

    def test_defaults_match_reference_constants(self):
        r = self._run()
        assert r.compatibility_rate == COMPATIBILITY_RATE
        assert r.bug_report_rate == BUG_REPORT_RATE
        assert r.pm_overhead == PM_OVERHEAD
        assert r.max_testing_stretch == MAX_TESTING_STRETCH

    def test_custom_pm_overhead_changes_pm_totals(self):
        base = self._run()
        custom = self._run(pm_overhead=0.20)
        assert custom.pm_overhead == 0.20
        # PM (one-platform) must scale linearly with the new overhead
        assert _close(custom.one_plat_pm_min,
                      base.one_plat_min * 0.20)
        assert _close(custom.one_plat_pm_max,
                      base.one_plat_max * 0.20)
        # Totals should rise accordingly
        assert custom.one_plat_total_min > base.one_plat_total_min

    def test_custom_bug_report_rate_feeds_bug_rows(self):
        custom = self._run(bug_report_rate=0.25)
        assert custom.bug_report_rate == 0.25
        func = next(t for t in custom.tasks if t.key == "functional")
        reg = next(t for t in custom.tasks if t.key == "regression")
        bug = next(t for t in custom.tasks if t.key == "bug_report")
        rec = next(t for t in custom.tasks if t.key == "bug_recheck")
        assert _close(bug.min_h, (func.min_h + reg.min_h) * 0.25)
        assert _close(bug.max_h, (func.max_h + reg.max_h) * 0.25)
        # Bug rechecking mirrors bug reporting
        assert _close(rec.min_h, bug.min_h)
        assert _close(rec.max_h, bug.max_h)

    def test_custom_compatibility_rate_feeds_compat_row(self):
        custom = self._run(compatibility_rate=0.01)
        assert custom.compatibility_rate == 0.01
        chk = next(t for t in custom.tasks if t.key == "checklist")
        com = next(t for t in custom.tasks if t.key == "compatibility")
        assert _close(com.min_h, chk.min_h * 0.01)
        assert _close(com.max_h, chk.max_h * 0.01)

    def test_custom_max_stretch_changes_checklist_max(self):
        custom = self._run(max_testing_stretch=2.0)
        assert custom.max_testing_stretch == 2.0
        chk = next(t for t in custom.tasks if t.key == "checklist")
        # MAX should now equal 2× MIN (instead of default 1.5×)
        assert _close(chk.max_h, 2.0 * chk.min_h)

    def test_all_four_coefficients_persist_on_result(self):
        r = self._run(compatibility_rate=0.005, bug_report_rate=0.18,
                      pm_overhead=0.10, max_testing_stretch=1.7)
        assert r.compatibility_rate == 0.005
        assert r.bug_report_rate == 0.18
        assert r.pm_overhead == 0.10
        assert r.max_testing_stretch == 1.7

    def test_team_size_drives_brooks_penalty(self):
        solo = self._run(team_size=1)
        team = self._run(team_size=4)
        assert solo.brooks_overhead_hours == 0.0
        assert team.brooks_overhead_hours > 0.0
        assert team.team_size == 4
        # PERT expected with team penalty should exceed the solo run.
        assert team.pert_expected > solo.pert_expected

    def test_custom_rates_flow_to_xlsx_formulas(self, tmp_path):
        """Exported XLSX must embed the user-supplied rates in every formula
        that references them (PM, bug rate, compatibility, MAX stretch)."""
        from openpyxl import load_workbook
        from engine.qa_estimator import export_estimation_xlsx

        r = self._run(compatibility_rate=0.005, bug_report_rate=0.20,
                      pm_overhead=0.12, max_testing_stretch=1.8)
        out = tmp_path / "custom_est.xlsx"
        export_estimation_xlsx(r, str(out))

        wb = load_workbook(str(out))
        ws = wb["Manual QA"]
        # Row 13 MAX formula uses stretch coefficient
        assert "1.8" in str(ws["F13"].value)
        # Bug reporting rows reference the bug rate
        assert "0.2" in str(ws["E15"].value)
        assert "0.2" in str(ws["F16"].value)
        # Compatibility rows reference the compat rate
        assert "0.005" in str(ws["E18"].value)
        # PM rows reference the PM overhead
        assert "0.12" in str(ws["E20"].value)
        assert "0.12" in str(ws["F24"].value)
