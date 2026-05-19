"""
Sprint 1 Task 7 — CSV/Excel formula injection sanitization.

Spreadsheet apps (Excel, LibreOffice, Google Sheets) interpret cells whose
text starts with `=`, `+`, `-`, `@`, `|`, `\\t`, or `\\r` as formulas. A
user-controlled summary or comment such as `=cmd|'/c calc'!A1` would
execute when the recipient opens the exported file. These tests verify
that engine.exporter prepends a single apostrophe (the canonical Excel
"treat as literal" escape) in front of any string cell beginning with one
of those characters — and that benign values are passed through unchanged.
"""

import csv
import io
import os
import sys

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from engine.exporter import (  # noqa: E402
    export_csv_checklist,
    export_csv_testcases,
    export_xlsx_checklist,
    export_xlsx_testcases,
)
from engine.testcase_generator import ChecklistItem, TestCase  # noqa: E402


# ── helpers ───────────────────────────────────────────────────────

def _make_tc(**overrides) -> TestCase:
    base = dict(
        id="SC1_001",
        section="Auth",
        section_num=1,
        summary="Login flow",
        preconditions="User has account",
        test_steps="1. Open page\n2. Enter credentials",
        test_data="user@example.com / Pass123",
        expected_result="User is logged in",
        issues="",
        comment="",
        user_story_id="US1",
        category="Positive",
        priority="High",
        status="Unchecked",
    )
    base.update(overrides)
    return TestCase(**base)


def _make_cl(**overrides) -> ChecklistItem:
    base = dict(
        id="HDR_001",
        section="Header",
        objective="Verify logo is visible",
        comments="",
        user_story_id="US1",
        category="Positive",
        priority="Medium",
        status="Unchecked",
    )
    base.update(overrides)
    return ChecklistItem(**base)


def _csv_rows(text: str) -> list[list[str]]:
    return list(csv.reader(io.StringIO(text)))


# ── CSV tests ─────────────────────────────────────────────────────

def test_csv_sanitizes_leading_equals():
    """A `=cmd|'/c calc'!A1` summary must be neutralized to start with `'=`."""
    payload = "=cmd|'/c calc'!A1"
    tc = _make_tc(summary=payload)
    csv_text = export_csv_testcases([tc])
    rows = _csv_rows(csv_text)
    # Header is row 0; data row is row 1. Summary is column index 2.
    summary_cell = rows[1][2]
    assert summary_cell.startswith("'="), (
        f"expected leading apostrophe before '=', got: {summary_cell!r}"
    )
    assert summary_cell == "'" + payload


def test_csv_sanitizes_at_pipe_tab():
    """All four remaining formula triggers (@, |, \\t, \\r) must be escaped."""
    cases = {
        "summary":         "@SUM(A1:A10)",   # @
        "preconditions":   "|pipe payload",  # |
        "test_data":       "\tTAB payload",  # \t
        "comment":         "\rCR payload",   # \r
    }
    tc = _make_tc(**cases)
    rows = _csv_rows(export_csv_testcases([tc]))
    header = rows[0]
    data = rows[1]

    col_index = {name: header.index(label) for name, label in {
        "summary":       "Summary",
        "preconditions": "Preconditions",
        "test_data":     "Test Data",
        "comment":       "Comment",
    }.items()}

    for field, original in cases.items():
        cell = data[col_index[field]]
        expected = "'" + original
        assert cell == expected, (
            f"{field} not sanitized; got {cell!r}, expected {expected!r}"
        )

    # Also verify the checklist CSV escapes the same triggers in `comments`.
    cl = _make_cl(objective="@evil()", comments="|payload")
    cl_rows = _csv_rows(export_csv_checklist([cl]))
    cl_header = cl_rows[0]
    obj_cell = cl_rows[1][cl_header.index("Objective")]
    com_cell = cl_rows[1][cl_header.index("Comments")]
    assert obj_cell == "'@evil()"
    assert com_cell == "'|payload"


# ── XLSX tests ────────────────────────────────────────────────────

def test_xlsx_cell_value_starts_with_apostrophe():
    """Open the produced XLSX with openpyxl and confirm the dangerous
    summary cell now begins with a literal apostrophe."""
    payload = "=cmd|'/c calc'!A1"
    tc = _make_tc(summary=payload)
    xlsx_bytes = export_xlsx_testcases([tc])

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    # Headers in row 1; Summary is column 3 in this exporter.
    headers = [c.value for c in ws[1]]
    summary_col = headers.index("Summary") + 1  # 1-based
    summary_cell_value = ws.cell(row=2, column=summary_col).value

    assert isinstance(summary_cell_value, str)
    assert summary_cell_value.startswith("'="), (
        f"XLSX summary cell not sanitized: {summary_cell_value!r}"
    )
    assert summary_cell_value == "'" + payload

    # Checklist XLSX should also escape leading `+` in the comments column.
    cl = _make_cl(comments="+danger")
    cl_bytes = export_xlsx_checklist([cl])
    cl_wb = load_workbook(io.BytesIO(cl_bytes))
    cl_ws = cl_wb.active
    cl_headers = [c.value for c in cl_ws[1]]
    com_col = cl_headers.index("Comments") + 1
    assert cl_ws.cell(row=2, column=com_col).value == "'+danger"


# ── Benign values must be untouched ───────────────────────────────

def test_benign_values_unchanged():
    """Normal text and numerics must pass through without modification."""
    tc = _make_tc(summary="Login flow", comment="No issues observed")
    rows = _csv_rows(export_csv_testcases([tc]))
    header = rows[0]
    data = rows[1]

    assert data[header.index("Summary")] == "Login flow"
    assert data[header.index("Comment")] == "No issues observed"
    # The benign priority "High" must remain "High", not "'High".
    assert data[header.index("Priority")] == "High"

    # XLSX path — same expectations.
    xlsx_bytes = export_xlsx_testcases([tc])
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    xlsx_headers = [c.value for c in ws[1]]
    summary_col = xlsx_headers.index("Summary") + 1
    priority_col = xlsx_headers.index("Priority") + 1
    assert ws.cell(row=2, column=summary_col).value == "Login flow"
    assert ws.cell(row=2, column=priority_col).value == "High"
