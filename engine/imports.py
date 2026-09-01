"""
TestFortge — import parsers for existing Test Cases and Checklists.

Tester teams already have test packs (TestRail / Zephyr / qase.io / TestFort
exports / hand-rolled spreadsheets). To run those packs in TestFortge —
manually via /test-execution or automated via /automation — the framework
needs a tolerant importer that accepts whichever shape the pack arrived in.

This module exposes two public functions:

    parse_test_cases(file_path, ext)  -> list[TestCase]
    parse_checklist(file_path, ext)   -> list[ChecklistItem]

Each accepts XLSX, CSV, MD, JSON. Header / column names are matched
case-insensitively against an alias dictionary so a column titled
"Test Steps", "Steps", "Procedure" or "Шаги" all map to the same
field. Missing optional columns get sensible defaults so a minimal
"ID + Summary + Steps + Expected" pack still imports cleanly.
"""
from __future__ import annotations

import csv
import io
import json
import os
import re
from typing import Any, Iterable

from openpyxl import load_workbook

from .log import get_logger
from .testcase_generator import TestCase, ChecklistItem

log = get_logger(__name__)


# ── Header alias maps ──────────────────────────────────────────────

# Each canonical key maps to a tuple of accepted header substrings.
# Match is case-insensitive and substring-based (after normalising
# whitespace and stripping non-alpha chars), so "TC ID", "TC-ID",
# "Test Case ID", "id", "ID#" all resolve to "id".
TC_ALIASES: dict[str, tuple[str, ...]] = {
    "id":               ("tcid", "id", "testcaseid", "caseid", "tc#", "#"),
    "section":          ("section", "module", "suite", "feature", "area"),
    "summary":          ("summary", "title", "testcase", "name", "description", "scenario"),
    "preconditions":    ("preconditions", "precondition", "preconds",
                         "setup", "given", "pre"),
    "test_steps":       ("teststeps", "steps", "procedure", "actions", "when",
                         "stepstoreproduce", "шаги", "крок"),
    "test_data":        ("testdata", "data", "inputs", "input"),
    "expected_result":  ("expectedresult", "expected", "passcriteria",
                         "expectedoutcome", "result", "then"),
    # ``category`` here means the quality flavour of the case
    # (Positive / Negative / Edge Case / Security) — distinct from
    # ``testing_type`` which is the discipline (Functional / SEO / etc.)
    "category":         ("category", "casetype", "scenariotype", "kind"),
    "priority":         ("priority", "severity", "importance"),
    "user_story_id":    ("userstory", "userstoryid", "storyid", "story"),
    "issues":           ("issues", "issue", "bug", "bugid", "linkedbug"),
    "comment":          ("comment", "comments", "notes", "note"),
    "status":           ("status", "execution", "executionstatus", "state"),
    "testing_type":     ("testingtype", "testtype", "type"),
}

CL_ALIASES: dict[str, tuple[str, ...]] = {
    "id":               ("id", "checkid", "checklistid", "#"),
    "section":          ("section", "module", "suite", "feature", "area"),
    "objective":        ("objective", "summary", "check", "description", "title", "item"),
    "category":         ("category", "casetype", "scenariotype", "kind"),
    "priority":         ("priority", "severity", "importance"),
    "user_story_id":    ("userstory", "userstoryid", "storyid"),
    "comments":         ("comments", "comment", "notes", "note"),
    "status":           ("status", "state"),
    "testing_type":     ("testingtype", "testtype", "type"),
}


# ── Header normalisation ───────────────────────────────────────────

def _norm(header: str) -> str:
    """Lower-case + strip everything but alphanum so 'TC-ID' == 'tcid'."""
    return re.sub(r"[^a-z0-9а-яіїєґ]", "", (header or "").lower())


def _build_header_map(headers: Iterable[str], aliases: dict[str, tuple[str, ...]]) -> dict[str, int]:
    """Map canonical-key → column-index by matching headers against aliases."""
    out: dict[str, int] = {}
    norms = [(_norm(h), i) for i, h in enumerate(headers)]
    for key, alias_tuple in aliases.items():
        # Try exact match first (highest precedence).
        for normed, idx in norms:
            if normed in alias_tuple:
                out[key] = idx
                break
        if key in out:
            continue
        # Substring match — looser fallback ("test_steps" matches "teststepstoreproduce").
        for normed, idx in norms:
            if any(a in normed or normed in a for a in alias_tuple):
                out[key] = idx
                break
    return out


# ── Row → record helpers ──────────────────────────────────────────

def _cell(row: list, col_map: dict[str, int], key: str, default: str = "") -> str:
    """Return row[col_map[key]] as a stripped string, with default fallback."""
    idx = col_map.get(key)
    if idx is None:
        return default
    val = row[idx] if idx < len(row) else None
    if val is None:
        return default
    return str(val).strip()


_ID_RE = re.compile(r"^[A-Za-z]{2,6}[_-]?\d{1,4}$|^SC\d+_\d{1,3}$|^TC[-_]?\d+$|^\d+$")


def _coerce_id(raw: str, fallback_prefix: str, idx: int) -> str:
    """Use the file's id if present, otherwise synthesise a stable one."""
    raw = (raw or "").strip()
    if raw and _ID_RE.match(raw):
        return raw
    if raw:
        # Some teams use long IDs like 'AUTH-001-PRIORITY' — keep them.
        return raw[:40]
    return f"{fallback_prefix}{idx:03d}"


def _row_to_test_case(row: list, col_map: dict[str, int], idx: int) -> TestCase | None:
    """Convert one row into a TestCase, or None if the row is empty."""
    summary = _cell(row, col_map, "summary")
    steps = _cell(row, col_map, "test_steps")
    if not summary and not steps:
        return None  # blank row
    section = _cell(row, col_map, "section") or "Imported"
    section_num = 1  # rebuilt by display layer; keep nominal
    tc_id = _coerce_id(_cell(row, col_map, "id"), "TC", idx)
    return TestCase(
        id=tc_id,
        section=section,
        section_num=section_num,
        summary=summary or f"Imported case {idx}",
        preconditions=_cell(row, col_map, "preconditions"),
        test_steps=steps,
        test_data=_cell(row, col_map, "test_data"),
        expected_result=_cell(row, col_map, "expected_result"),
        issues=_cell(row, col_map, "issues"),
        comment=_cell(row, col_map, "comment"),
        user_story_id=_cell(row, col_map, "user_story_id"),
        category=_cell(row, col_map, "category") or "Positive",
        priority=_cell(row, col_map, "priority") or "Medium",
        status=_cell(row, col_map, "status") or "Unchecked",
        testing_type=_cell(row, col_map, "testing_type") or "Functional",
    )


def _row_to_checklist_item(row: list, col_map: dict[str, int], idx: int) -> ChecklistItem | None:
    objective = _cell(row, col_map, "objective")
    if not objective:
        return None
    section = _cell(row, col_map, "section") or "Imported"
    cl_id = _coerce_id(_cell(row, col_map, "id"), "CL", idx)
    return ChecklistItem(
        id=cl_id,
        section=section,
        objective=objective,
        comments=_cell(row, col_map, "comments"),
        user_story_id=_cell(row, col_map, "user_story_id"),
        category=_cell(row, col_map, "category") or "Positive",
        priority=_cell(row, col_map, "priority") or "Medium",
        status=_cell(row, col_map, "status") or "Unchecked",
        testing_type=_cell(row, col_map, "testing_type") or "Functional",
    )


# ── Format-specific readers ───────────────────────────────────────

class ImportTooLarge(ValueError):
    """More rows than this instance will parse in one request.

    A ``ValueError`` so the upload routes' existing ``except Exception`` turns
    it into the flash they already show — and the message names the cap and
    what to do, because "import failed" without a number is not a diagnosis.

    **Refused rather than truncated**, deliberately. Half a test pack that
    looks complete is worse than an upload that did not happen: the operator
    would have to notice the missing rows themselves, and cases that silently
    vanished are indistinguishable from cases nobody wrote.
    """


#: Rows one upload may carry. Bounded because the reader below is fed
#: directly by a browser upload and nothing else bounded it.
#:
#: Measured on a **2.5 MB** file of 200 000 rows, which ``MAX_CONTENT_LENGTH``
#: (64 MB) would accept twenty-five times over:
#:
#:   * ``load_workbook`` + ``list(iter_rows())`` — 175 MB peak, 56 s;
#:   * ``read_only=True``, streamed — 16 MB, 47 s;
#:   * ``read_only=True`` with this cap — 2 MB, 5.8 s;
#:   * and :func:`read_headers`, which wants *one row*, paid 190 MB and 75 s,
#:     because it went through the same full parse.
#:
#: The upload route calls ``read_headers`` and then parses again, so that one
#: file cost roughly 365 MB of peak allocation and over two minutes — on a
#: 512 MB dyno running ``--workers 1`` behind a proxy that closes idle
#: connections at 30 s. The worker dies and takes every other request with it.
#:
#: 20 000 against a measured reference plan of 4 808 cases: four times the
#: largest real pack this project has seen.
MAX_IMPORT_ROWS = int(os.environ.get("MAX_IMPORT_ROWS", "20000"))


def _too_large() -> ImportTooLarge:
    return ImportTooLarge(
        f"this file has more than {MAX_IMPORT_ROWS:,} rows. Split it and "
        f"upload the parts, or raise MAX_IMPORT_ROWS on the service.")


def _read_xlsx(file_path: str) -> tuple[list[str], list[list]]:
    """Return (header_row, [data_rows]) from the active sheet of an XLSX file.

    ``read_only=True`` streams the sheet instead of building every cell
    object first — 175 MB against 16 MB on the file measured above. It also
    yields only the rows that exist, so an inflated ``<dimension>`` costs
    nothing; checked rather than assumed, which is what lets the cap count
    yielded rows without worrying about padding.
    """
    wb = load_workbook(file_path, data_only=True, read_only=True)
    try:
        header: list[str] = []
        data: list[list] = []
        for index, row in enumerate(wb.active.iter_rows(values_only=True)):
            if index == 0:
                header = [str(c) if c is not None else "" for c in row]
                continue
            if not any(c is not None and str(c).strip() for c in row):
                continue
            if len(data) >= MAX_IMPORT_ROWS:
                raise _too_large()
            data.append(list(row))
    finally:
        wb.close()
    if not header:
        return ([], [])
    return (header, data)


def _read_xlsx_header(file_path: str) -> list[str]:
    """The header row alone, without parsing the sheet behind it.

    :func:`read_headers` called :func:`_read_xlsx` and threw away everything
    but ``[0]``. On the measured file that was 190 MB and 75 s to learn two
    column names; this is 0.7 MB and 0.02 s.
    """
    wb = load_workbook(file_path, data_only=True, read_only=True)
    try:
        first = next(iter(wb.active.iter_rows(values_only=True)), None)
    finally:
        wb.close()
    if first is None:
        return []
    return [str(c) if c is not None else "" for c in first]


def _read_csv(file_path: str) -> tuple[list[str], list[list]]:
    """Read a CSV file — auto-detect delimiter (comma / semicolon / tab)."""
    with open(file_path, "r", encoding="utf-8-sig", errors="replace") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            delim = dialect.delimiter
        except csv.Error:
            delim = ","
        rows: list[list] = []
        for row in csv.reader(f, delimiter=delim):
            if not any((c or "").strip() for c in row):
                continue
            if len(rows) > MAX_IMPORT_ROWS:      # the header plus the cap
                raise _too_large()
            rows.append(row)
    if not rows:
        return ([], [])
    return (rows[0], rows[1:])


def _read_json(file_path: str) -> list[dict]:
    """Read a TestFortge-style JSON export. Accepts either a list of dicts
    at the top level or {"test_cases": [...], "checklist": [...]} envelope."""
    with open(file_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for key in ("test_cases", "testCases", "checklist", "items", "cases"):
            if isinstance(data.get(key), list):
                return data[key]
    return []


# ── Markdown reader ────────────────────────────────────────────────
#
# We support two flavours, matching what TestFortge itself exports:
#
# 1) "Field-list" flavour (TestFortge MD export):
#       ## TC-001 — Verify login with valid creds
#       - **Section:** Authentication
#       - **Preconditions:** ...
#       - **Test Steps:** 1. Open login. 2. ...
#       - **Test Data:** ...
#       - **Expected Result:** ...
#       - **Category:** Positive
#       - **Priority:** High
#
# 2) "Bullet-list" flavour (lightweight checklist):
#       ## Section name
#       - Verify that ...
#       - Verify that ...

_MD_HEADING_RE = re.compile(r"^(#{1,4})\s+(.+?)\s*$")
_MD_TC_TITLE_RE = re.compile(r"^([A-Za-z0-9_\-]+)\s*[—\-:]\s*(.+)$")
_MD_FIELD_RE = re.compile(r"^\s*[-*]\s*\*\*([A-Za-z][A-Za-z _/]+):\*\*\s*(.+)$",
                           re.IGNORECASE)
_MD_BULLET_RE = re.compile(r"^\s*[-*]\s+(.+)$")


def _read_md_test_cases(file_path: str) -> list[TestCase]:
    text = open(file_path, "r", encoding="utf-8", errors="replace").read()
    lines = text.split("\n")
    cases: list[TestCase] = []
    current: dict[str, str] = {}
    current_section = ""
    next_idx = 1

    def _flush():
        nonlocal next_idx
        if current.get("summary"):
            tc_id = current.get("id") or f"TC{next_idx:03d}"
            cases.append(TestCase(
                id=tc_id, section=current.get("section") or current_section or "Imported",
                section_num=1,
                summary=current["summary"],
                preconditions=current.get("preconditions", ""),
                test_steps=current.get("test_steps", ""),
                test_data=current.get("test_data", ""),
                expected_result=current.get("expected_result", ""),
                issues=current.get("issues", ""),
                comment=current.get("comment", ""),
                user_story_id=current.get("user_story_id", ""),
                category=current.get("category", "Positive"),
                priority=current.get("priority", "Medium"),
                status=current.get("status", "Unchecked"),
                testing_type=current.get("testing_type", "Functional"),
            ))
            next_idx += 1
        current.clear()

    for raw in lines:
        h = _MD_HEADING_RE.match(raw)
        if h:
            level, title = len(h.group(1)), h.group(2).strip()
            if level <= 2:
                _flush()
                current_section = title
                continue
            # heading is a TC anchor "## TC-001 — Summary"
            _flush()
            m_title = _MD_TC_TITLE_RE.match(title)
            if m_title:
                current["id"] = m_title.group(1).strip()
                current["summary"] = m_title.group(2).strip()
            else:
                current["summary"] = title
            continue

        m_field = _MD_FIELD_RE.match(raw)
        if m_field and current.get("summary"):
            label = m_field.group(1).strip().lower().replace(" ", "_").replace("/", "_")
            value = m_field.group(2).strip()
            field_aliases = {
                "test_steps": "test_steps",
                "steps": "test_steps",
                "preconditions": "preconditions",
                "test_data": "test_data",
                "data": "test_data",
                "expected_result": "expected_result",
                "expected": "expected_result",
                "category": "category",
                "priority": "priority",
                "section": "section",
                "user_story": "user_story_id",
                "story": "user_story_id",
                "testing_type": "testing_type",
                "type": "testing_type",
            }
            current[field_aliases.get(label, label)] = value
            continue

        m_bul = _MD_BULLET_RE.match(raw)
        if m_bul and not current.get("summary") and current_section:
            # Lightweight bulleted MD: each bullet is a separate TC.
            line = m_bul.group(1).strip()
            cases.append(TestCase(
                id=f"TC{next_idx:03d}", section=current_section,
                section_num=1, summary=line,
                preconditions="", test_steps="", test_data="",
                expected_result="", issues="", comment="",
                user_story_id="", category="Positive", priority="Medium",
                status="Unchecked", testing_type="Functional",
            ))
            next_idx += 1

    _flush()
    return cases


def _read_md_checklist(file_path: str) -> list[ChecklistItem]:
    text = open(file_path, "r", encoding="utf-8", errors="replace").read()
    lines = text.split("\n")
    items: list[ChecklistItem] = []
    current_section = ""
    next_idx = 1

    for raw in lines:
        h = _MD_HEADING_RE.match(raw)
        if h:
            current_section = h.group(2).strip()
            continue
        m = _MD_BULLET_RE.match(raw)
        if not m:
            continue
        text_line = m.group(1).strip()
        # Strip leading "[ ]" / "[x]" task markers if present
        text_line = re.sub(r"^\[[\sxX]\]\s*", "", text_line)
        items.append(ChecklistItem(
            id=f"CL{next_idx:03d}",
            section=current_section or "Imported",
            objective=text_line,
            comments="", user_story_id="",
            category="Positive", priority="Medium",
            status="Unchecked", testing_type="Functional",
        ))
        next_idx += 1
    return items


# ── Public API ─────────────────────────────────────────────────────

def read_headers(file_path: str, filename: str = "") -> list[str]:
    """The file's own column names, for the mapping form (E4.8).

    Empty for the formats that have no header row (Markdown, JSON) — those
    carry their field names inside each record, so there is nothing to map.
    """
    ext = (filename or os.path.basename(file_path)).rsplit(".", 1)[-1].lower()
    try:
        if ext == "xlsx":
            return list(_read_xlsx_header(file_path) or [])
        if ext == "csv":
            return list(_read_csv(file_path)[0] or [])
        if ext == "json":
            records = _read_json(file_path)
            return list(records[0].keys()) if records else []
    except Exception as exc:      # pragma: no cover — unreadable file
        log.warning("could not read headers from %s: %s", filename, exc)
    return []


def _column_map(headers, aliases, mapping=None):
    """The automatic match, with an explicit user mapping laid over it."""
    from engine import import_preview
    base = _build_header_map(headers, aliases)
    if not mapping:
        return base
    kind = "test_cases" if aliases is TC_ALIASES else "checklist"
    return import_preview.resolve(kind, headers, mapping, base=base)


def parse_test_cases(file_path: str, filename: str = "",
                     mapping: dict | None = None) -> list[TestCase]:
    """Parse uploaded file → list[TestCase]. Format inferred from extension."""
    ext = (filename or os.path.basename(file_path)).rsplit(".", 1)[-1].lower()
    if ext == "xlsx":
        header, rows = _read_xlsx(file_path)
        if not header:
            return []
        col_map = _column_map(header, TC_ALIASES, mapping)
        out = []
        for i, row in enumerate(rows, start=1):
            tc = _row_to_test_case(row, col_map, i)
            if tc:
                out.append(tc)
        return out
    if ext == "csv":
        header, rows = _read_csv(file_path)
        if not header:
            return []
        col_map = _column_map(header, TC_ALIASES, mapping)
        out = []
        for i, row in enumerate(rows, start=1):
            tc = _row_to_test_case(row, col_map, i)
            if tc:
                out.append(tc)
        return out
    if ext in ("md", "markdown"):
        return _read_md_test_cases(file_path)
    if ext == "json":
        records = _read_json(file_path)
        out = []
        for i, rec in enumerate(records, start=1):
            # Reuse row-level conversion by faking a single-row table.
            keys = list(rec.keys())
            col_map = _column_map(keys, TC_ALIASES, mapping)
            row = [rec.get(k) for k in keys]
            tc = _row_to_test_case(row, col_map, i)
            if tc:
                out.append(tc)
        return out
    raise ValueError(f"Unsupported file extension for test cases: {ext!r}. "
                     "Supported: xlsx, csv, md, json.")


def parse_checklist(file_path: str, filename: str = "",
                    mapping: dict | None = None) -> list[ChecklistItem]:
    """Parse uploaded file → list[ChecklistItem]."""
    ext = (filename or os.path.basename(file_path)).rsplit(".", 1)[-1].lower()
    if ext == "xlsx":
        header, rows = _read_xlsx(file_path)
        if not header:
            return []
        col_map = _column_map(header, CL_ALIASES, mapping)
        out = []
        for i, row in enumerate(rows, start=1):
            cl = _row_to_checklist_item(row, col_map, i)
            if cl:
                out.append(cl)
        return out
    if ext == "csv":
        header, rows = _read_csv(file_path)
        if not header:
            return []
        col_map = _column_map(header, CL_ALIASES, mapping)
        out = []
        for i, row in enumerate(rows, start=1):
            cl = _row_to_checklist_item(row, col_map, i)
            if cl:
                out.append(cl)
        return out
    if ext in ("md", "markdown"):
        return _read_md_checklist(file_path)
    if ext == "json":
        records = _read_json(file_path)
        out = []
        for i, rec in enumerate(records, start=1):
            keys = list(rec.keys())
            col_map = _column_map(keys, CL_ALIASES, mapping)
            row = [rec.get(k) for k in keys]
            cl = _row_to_checklist_item(row, col_map, i)
            if cl:
                out.append(cl)
        return out
    raise ValueError(f"Unsupported file extension for checklist: {ext!r}. "
                     "Supported: xlsx, csv, md, json.")
