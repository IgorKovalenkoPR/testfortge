"""
TestFortge — QA Effort Estimator

Authored by the QA Team Lead persona (ISTQB Advanced, 10+ years across
FinTech, E-commerce, Healthcare, SaaS, EdTech, Telecom, Government).

Produces Manual QA/Testing effort estimations that mirror the reference
template `Example of Manual QA.Estimation.v1.0.xlsx`.

Formulas (match the reference exactly):

  Features sheet
    total_tc  = SUM(test_cases per feature)
    hours_tc  = int(buffer * total_tc * minutes_per_tc / 60) + 1      # default buffer=1.12, minutes=5

  Manual QA sheet (E = MIN hrs, F = MAX hrs, G = Expected = AVG(E, F))
    11. Communication               : E=12,  F=16
    12. Testing docs & requirements : E=12,  F=16
    13. Checklist Creation          : E = hours_tc,           F = 1.5 * E
    14. Functional & UI Testing     : E = E13,                F = F13
    15. Bug Reporting               : E = (E14 + E17) * 0.15, F = (F14 + F17) * 0.15
    16. Bug Rechecking              : E = (E14 + E17) * 0.15, F = (F14 + F17) * 0.15
    17. Regression Testing          : E = E14,                F = F14,   G = G14
    18. Compatibility (per extra)   : E = E13 * 0.003,        F = F13 * 0.003

  Totals (One Platform)
    19. Testing/QA     = SUM(E11..E17), SUM(F11..F17)
    20. PM             = 19 * 0.08
    21. Total          = 19 + 20

  Totals (Full Compatibility, N extra platforms)
    23. Testing/QA     = SUM(E11..E17) + E18 * N
    24. PM             = 23 * 0.08
    25. Total          = 23 + 24

  Cost
    28. Cost w/o compatibility  = Total (21) * rate
    29. Cost with compatibility = Total (25) * rate
"""
from __future__ import annotations

import math
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Iterable

from engine.log import get_logger

_logger = get_logger(__name__)


# ── Data structures ──────────────────────────────────────────────

@dataclass
class Feature:
    """A single module/page/feature with estimated test-case count."""
    name: str
    test_cases: int = 0
    comment: str = ""
    is_section: bool = False          # True for grouping headers with no own tests


@dataclass
class TaskRow:
    """A single row in the estimation table.

    Carries both the legacy two-point fields (min_h / max_h /
    expected_h = avg) AND a triangular / Beta-PERT layer
    (most_likely_h, pert_expected_h, sigma_h). The XLSX export and the
    classic UI continue to read the legacy fields unchanged; the new
    fields are surfaced by the UI as a confidence panel without
    altering any client-facing deliverable.
    """
    key: str           # stable identifier (e.g. "communication")
    title: str         # human-readable title
    description: str   # long description shown in the sheet
    min_h: float
    max_h: float
    expected_h: float
    # ── Triangular / Beta-PERT (additive — never replaces expected_h)
    most_likely_h: float = 0.0      # M — heuristic between O and P
    pert_expected_h: float = 0.0    # (O + 4M + P) / 6
    sigma_h: float = 0.0            # (P - O) / 6


@dataclass
class EstimationResult:
    project_name: str = ""
    rate_usd: float = 0.0
    additional_platforms: int = 9
    primary_platform: str = "Windows 10"
    platforms_list: list[str] = field(default_factory=list)
    minutes_per_tc: int = 5
    buffer: float = 1.12
    # Per-run rates (editable by the user; defaults match the reference template)
    compatibility_rate: float = 0.003
    bug_report_rate: float = 0.15
    pm_overhead: float = 0.08
    max_testing_stretch: float = 1.5
    source: str = ""                       # "url" | "attachment" | "manual"
    source_ref: str = ""                   # URL or file name
    created_at: str = ""

    features: list[Feature] = field(default_factory=list)
    total_tc: int = 0
    features_hours: int = 0                # Features!B40
    # Heuristic complexity bucket — drives Communication / Testing-doc hours.
    # One of: "simple" | "medium" | "complex".
    complexity_tier: str = "simple"

    tasks: list[TaskRow] = field(default_factory=list)

    # One Platform totals
    one_plat_min: float = 0.0
    one_plat_max: float = 0.0
    one_plat_expected: float = 0.0
    one_plat_pm_min: float = 0.0
    one_plat_pm_max: float = 0.0
    one_plat_pm_expected: float = 0.0
    one_plat_total_min: float = 0.0
    one_plat_total_max: float = 0.0
    one_plat_total_expected: float = 0.0

    # Full Compatibility totals
    full_min: float = 0.0
    full_max: float = 0.0
    full_expected: float = 0.0
    full_pm_min: float = 0.0
    full_pm_max: float = 0.0
    full_pm_expected: float = 0.0
    full_total_min: float = 0.0
    full_total_max: float = 0.0
    full_total_expected: float = 0.0

    # Costs (USD)
    cost_one_min: float = 0.0
    cost_one_max: float = 0.0
    cost_one_expected: float = 0.0
    cost_full_min: float = 0.0
    cost_full_max: float = 0.0
    cost_full_expected: float = 0.0

    # ── Beta-PERT aggregates (additive layer, never alters legacy fields above)
    # Aggregated across the 7 core task rows. ``pert_sigma`` is the
    # combined standard deviation under the assumption that task durations
    # are independent — sqrt(sum of variances). Confidence bands follow
    # the 68-95-99.7 rule for a normal approximation of Beta-PERT.
    team_size: int = 1
    # System-suggested team size derived from the computed effort and
    # complexity tier. Shown in the UI as a "Suggested: N" badge — the
    # user can still Override to a manual value before re-running the
    # estimation. Default 1 keeps legacy behaviour for callers that
    # don't recompute it after construction.
    suggested_team_size: int = 1
    pert_expected: float = 0.0
    pert_sigma: float = 0.0
    band_68_low: float = 0.0
    band_68_high: float = 0.0
    band_95_low: float = 0.0
    band_95_high: float = 0.0
    band_99_low: float = 0.0
    band_99_high: float = 0.0
    # Brooks's-Law communication-overhead penalty in hours (0 when
    # team_size <= 1). Surfaced in the UI as its own line so the user
    # sees why the expected-with-team grew — XLSX rows untouched.
    brooks_overhead_hours: float = 0.0
    # Historical calibration — populated by the route layer after a
    # peer-group lookup. When set, the UI shows a soft hint explaining
    # how the new estimate compares to past projects of the same owner
    # and similar feature count. Never overrides the actual numbers.
    history_hint: str = ""
    history_median_hours_per_feature: float = 0.0
    history_sample_size: int = 0


# ── Core formulas ────────────────────────────────────────────────

PM_OVERHEAD = 0.08          # Project Management share
BUG_REPORT_RATE = 0.15      # (Functional + Regression) * 0.15
COMPATIBILITY_RATE = 0.003  # 0.3% of Checklist creation hours per extra platform
MAX_TESTING_STRETCH = 1.5   # MAX = 1.5 * MIN for checklist/functional


# Communication + Testing-documentation hours scale by project complexity.
# Bands chosen by the QA Team Lead persona on real engagements:
#   simple  — landing page / brochure site, 1 platform, single role
#   medium  — typical SMB SaaS / e-commerce, 2-3 roles, a couple of platforms
#   complex — multi-tenant / fintech / multi-platform with admin + B2C + API
# These cover both Communication (row 11) and Testing documentation &
# requirements (row 12) per the reference template.
_COMM_DOC_HOURS: dict[str, tuple[float, float]] = {
    "simple":  (6.0,  8.0),
    "medium":  (9.0, 11.0),
    "complex": (12.0, 16.0),
}


def _complexity_tier(total_tc: int, features_count: int,
                     additional_platforms: int) -> str:
    """Heuristic complexity bucket — 'simple', 'medium' or 'complex'.

    Combines three signals because no single one is reliable on its own:
      * test-case count       — primary proxy for surface area
      * feature count         — domain breadth / number of flows
      * additional platforms  — multi-platform coverage adds coordination

    Each signal contributes points; the sum maps to a bucket. Thresholds
    were calibrated against the reference Manual QA template so that:
      ≤ 50 TC, ≤ 5 features, ≤ 1 platform → simple
      ~ 100 TC, ~ 10 features              → medium
      ≥ 200 TC OR many platforms           → complex
    """
    score = 0.0

    # Test-case surface area (dominant signal)
    if   total_tc <= 50:  score += 0.0
    elif total_tc <= 120: score += 1.0
    elif total_tc <= 250: score += 2.0
    else:                 score += 3.0

    # Feature breadth
    if   features_count <= 4:  score += 0.0
    elif features_count <= 12: score += 0.5
    else:                      score += 1.0

    # Multi-platform coordination overhead
    if   additional_platforms <= 1: score += 0.0
    elif additional_platforms <= 5: score += 0.25
    else:                           score += 0.5

    if score <= 1.0:  return "simple"
    if score <= 2.5:  return "medium"
    return "complex"


def compute_features_hours(total_tc: int, minutes_per_tc: int = 5,
                            buffer: float = 1.12) -> int:
    """Features!B40 = int(buffer * total_tc * minutes_per_tc / 60) + 1.

    Returns 0 when there are no test cases, so the whole estimation reads as
    zero until the user supplies real input.
    """
    if not total_tc or total_tc <= 0:
        return 0
    return int(buffer * total_tc * minutes_per_tc / 60) + 1


def _avg(a: float, b: float) -> float:
    return (a + b) / 2.0


def suggest_team_size(total_hours: float, complexity_tier: str = "simple") -> int:
    """Recommend a tester headcount from the computed effort and tier.

    The heuristic mirrors how a QA Lead sizes a team in practice: small
    projects get a solo tester, mid-sized projects a pair, and only
    multi-hundred-hour engagements warrant 3+. Adds +1 for the "complex"
    tier so payment/security-heavy work isn't shoehorned into the same
    headcount as a brochure site of identical hour count. Hard-capped
    at 12 so absurd inputs don't break the UI badge.
    """
    h = float(total_hours or 0)
    if h < 50:
        base = 1
    elif h < 200:
        base = 2
    elif h < 500:
        base = 3
    elif h < 1000:
        base = 5
    else:
        base = 7
    if complexity_tier == "complex":
        base += 1
    return min(base, 12)


def compute_estimation(
    features: list[Feature],
    rate_usd: float = 0.0,
    additional_platforms: int = 9,
    minutes_per_tc: int = 5,
    buffer: float = 1.12,
    project_name: str = "",
    primary_platform: str = "Windows 10",
    platforms_list: Iterable[str] | None = None,
    source: str = "manual",
    source_ref: str = "",
    # Per-run, user-tunable coefficients. Defaults fall back to the
    # module-level reference values so existing call sites stay correct.
    compatibility_rate: float | None = None,
    bug_report_rate: float | None = None,
    pm_overhead: float | None = None,
    max_testing_stretch: float | None = None,
    # Team size for Brooks's-Law communication overhead. Default=1 keeps
    # legacy behaviour byte-identical for callers that don't pass it.
    team_size: int = 1,
) -> EstimationResult:
    """Apply the estimation formulas from the reference template.

    Every coefficient can be overridden per call — the QA Team Lead or the
    user can supply custom rates for compatibility, bug-report share, PM
    overhead and the MIN→MAX stretch. When a parameter is omitted the
    module-level default (matching the reference XLSX) is used.
    """
    _compat = COMPATIBILITY_RATE if compatibility_rate is None else float(compatibility_rate)
    _bug    = BUG_REPORT_RATE    if bug_report_rate    is None else float(bug_report_rate)
    _pm     = PM_OVERHEAD        if pm_overhead        is None else float(pm_overhead)
    _stretch = MAX_TESTING_STRETCH if max_testing_stretch is None else float(max_testing_stretch)

    res = EstimationResult(
        project_name=project_name,
        rate_usd=float(rate_usd or 0.0),
        additional_platforms=int(additional_platforms),
        primary_platform=primary_platform,
        platforms_list=list(platforms_list or []),
        minutes_per_tc=int(minutes_per_tc),
        buffer=float(buffer),
        compatibility_rate=_compat,
        bug_report_rate=_bug,
        pm_overhead=_pm,
        max_testing_stretch=_stretch,
        team_size=max(1, int(team_size or 1)),
        source=source,
        source_ref=source_ref,
        created_at=datetime.now().strftime("%Y-%m-%d"),
        features=list(features),
    )

    # Features totals
    res.total_tc = sum(f.test_cases for f in features if not f.is_section)
    res.features_hours = compute_features_hours(res.total_tc, minutes_per_tc, buffer)

    # All task hours collapse to 0 when there are no test cases, so that the
    # "Estimation Breakdown" reads as zero by default and only gets populated
    # once the user supplies features for a concrete project.
    has_tc = res.total_tc > 0
    # Communication and documentation hours scale with project complexity —
    # see ``_complexity_tier`` below for the heuristic. A simple landing
    # page should not get the same 12/16 hours as a multi-platform fintech.
    tier = _complexity_tier(res.total_tc,
                            sum(1 for f in features if not f.is_section),
                            int(additional_platforms))
    comm_min, comm_max = _COMM_DOC_HOURS[tier] if has_tc else (0.0, 0.0)
    doc_min,  doc_max  = _COMM_DOC_HOURS[tier] if has_tc else (0.0, 0.0)
    res.complexity_tier = tier
    # Row 13 Checklist Creation
    chk_min = float(res.features_hours)
    chk_max = _stretch * chk_min
    # Row 14 Functional & UI Testing
    func_min, func_max = chk_min, chk_max
    # Row 17 Regression Testing
    reg_min, reg_max = func_min, func_max
    # Row 15 Bug Reporting and Row 16 Bug Rechecking
    bug_min = (func_min + reg_min) * _bug
    bug_max = (func_max + reg_max) * _bug
    recheck_min, recheck_max = bug_min, bug_max
    # Row 18 Compatibility Testing (per each extra platform)
    compat_min = chk_min * _compat
    compat_max = chk_max * _compat

    res.tasks = [
        TaskRow("communication", "Communication",
                "Communications with customer during the project, meetings and discussions.",
                comm_min, comm_max, _avg(comm_min, comm_max)),
        TaskRow("documentation", "Testing documentation and requirements",
                "Go through the requirements document to ensure it is full, testable and consistent; "
                "also includes time needed to become familiar with the application under test.",
                doc_min, doc_max, _avg(doc_min, doc_max)),
        TaskRow("checklist", "Checklist Creation",
                f"Create a checklist/test suite covering all documented scenarios "
                f"({res.total_tc} estimated tests). The number of cases may grow after a broader "
                f"understanding of the application.",
                chk_min, chk_max, _avg(chk_min, chk_max)),
        TaskRow("functional", "Functional and UI Testing",
                "Manual execution of positive and negative test cases to discover defects and "
                "improvement areas.",
                func_min, func_max, _avg(func_min, func_max)),
        TaskRow("bug_report", "Bugs Reporting",
                "All bugs, errors and suggestions found during execution are reported and delivered.",
                bug_min, bug_max, _avg(bug_min, bug_max)),
        TaskRow("bug_recheck", "Bug Rechecking",
                "Retest the application to guarantee that fixed issues no longer occur and do not "
                "introduce regressions.",
                recheck_min, recheck_max, _avg(recheck_min, recheck_max)),
        TaskRow("regression", "Regression Testing",
                "Check that all features continue to work after the bug-fix phase and code freeze.",
                reg_min, reg_max, _avg(reg_min, reg_max)),  # expected = G14 per reference
        TaskRow("compatibility", "Compatibility Testing (per extra platform)",
                "Positive-path compatibility run on an additional platform combination.",
                compat_min, compat_max, _avg(compat_min, compat_max)),
    ]

    # One Platform: SUM(rows 11..17) — first 7 tasks, excluding compatibility (index 7)
    core = res.tasks[:7]
    res.one_plat_min = sum(t.min_h for t in core)
    res.one_plat_max = sum(t.max_h for t in core)
    res.one_plat_expected = sum(t.expected_h for t in core)
    res.one_plat_pm_min = res.one_plat_min * _pm
    res.one_plat_pm_max = res.one_plat_max * _pm
    res.one_plat_pm_expected = res.one_plat_expected * _pm
    res.one_plat_total_min = res.one_plat_min + res.one_plat_pm_min
    res.one_plat_total_max = res.one_plat_max + res.one_plat_pm_max
    res.one_plat_total_expected = res.one_plat_expected + res.one_plat_pm_expected

    # System suggestion for tester headcount, derived from one-platform
    # expected hours and the complexity tier. Stored alongside the
    # user-set team_size so the UI can render "Suggested: N" while still
    # accepting an override on the next form submit.
    res.suggested_team_size = suggest_team_size(
        res.one_plat_total_expected, res.complexity_tier,
    )

    # Full Compatibility: SUM(rows 11..17) + compatibility * N
    n = res.additional_platforms
    res.full_min = res.one_plat_min + compat_min * n
    res.full_max = res.one_plat_max + compat_max * n
    res.full_expected = _avg(res.full_min, res.full_max)
    res.full_pm_min = res.full_min * _pm
    res.full_pm_max = res.full_max * _pm
    res.full_pm_expected = res.full_expected * _pm
    res.full_total_min = res.full_min + res.full_pm_min
    res.full_total_max = res.full_max + res.full_pm_max
    res.full_total_expected = res.full_expected + res.full_pm_expected

    # Costs
    rate = res.rate_usd
    res.cost_one_min = res.one_plat_total_min * rate
    res.cost_one_max = res.one_plat_total_max * rate
    res.cost_one_expected = res.one_plat_total_expected * rate
    res.cost_full_min = res.full_total_min * rate
    res.cost_full_max = res.full_total_max * rate
    res.cost_full_expected = res.full_total_expected * rate

    # ── Beta-PERT layer (additive) ───────────────────────────────
    # Per-row M (most-likely) heuristic: nudge ~40% above the optimistic
    # floor — empirically matches QA effort distributions where the floor
    # is the "happy path" runtime and the ceiling builds in unknowns.
    # Source: standard PERT practice + Wikipedia 'Three-point estimation'.
    import math as _math
    pert_total_expected = 0.0
    pert_total_variance = 0.0
    for t in res.tasks[:7]:  # 7 core rows only — compatibility scales separately
        O, P = float(t.min_h), float(t.max_h)
        M = O + 0.4 * (P - O)
        sigma = (P - O) / 6.0
        e_pert = (O + 4.0 * M + P) / 6.0
        t.most_likely_h = round(M, 3)
        t.pert_expected_h = round(e_pert, 3)
        t.sigma_h = round(sigma, 3)
        pert_total_expected += e_pert
        pert_total_variance += sigma * sigma
    pert_total_sigma = _math.sqrt(pert_total_variance)
    # PM overhead is multiplicative on top of the core 7 rows.
    pert_total_expected_with_pm = pert_total_expected * (1.0 + _pm)
    pert_total_sigma_with_pm    = pert_total_sigma * (1.0 + _pm)

    # ── Brooks's Law: communication overhead with team_size > 1 ─
    # Each engineer past the first adds a small amount of comm overhead.
    # We use 7% per additional channel (n*(n-1)/2), capped at +35% so a
    # very large team doesn't inflate to absurd numbers without a human
    # in the loop. team_size <= 1 → zero penalty.
    n = max(1, int(res.team_size or 1))
    if n > 1:
        channels = n * (n - 1) / 2
        brooks_factor = min(0.35, 0.07 * channels)
        res.brooks_overhead_hours = round(
            pert_total_expected_with_pm * brooks_factor, 2,
        )
    else:
        res.brooks_overhead_hours = 0.0

    res.pert_expected = round(
        pert_total_expected_with_pm + res.brooks_overhead_hours, 2,
    )
    res.pert_sigma = round(pert_total_sigma_with_pm, 2)
    # 68-95-99.7 confidence bands assuming a normal approximation
    # of the Beta-PERT distribution. Floor at 0 — negative hours are nonsense.
    res.band_68_low  = round(max(0.0, res.pert_expected - res.pert_sigma), 2)
    res.band_68_high = round(res.pert_expected + res.pert_sigma, 2)
    res.band_95_low  = round(max(0.0, res.pert_expected - 2 * res.pert_sigma), 2)
    res.band_95_high = round(res.pert_expected + 2 * res.pert_sigma, 2)
    res.band_99_low  = round(max(0.0, res.pert_expected - 3 * res.pert_sigma), 2)
    res.band_99_high = round(res.pert_expected + 3 * res.pert_sigma, 2)

    return res


# ── Feature extraction ───────────────────────────────────────────

# A line that looks like a numbered/bulleted feature item:
#   "- Something", "* Something", "1. Something", "1) Something"
_ITEM_RE = re.compile(r"^\s*(?:[-*•●—–]|\d+[.)])\s+(.{2,150})$")

# Tabular line as produced by file_parser._parse_xlsx / _parse_csv:
#   "Banners/Advertising | 55 | Includes negative scenarios"
_TABLE_RE = re.compile(r"^(.+?)\s*\|\s*(\d{1,4})\s*(?:\|\s*(.*))?$")

# Keywords that hint a feature/module/page
_FEATURE_HINTS = re.compile(
    r"\b(module|feature|page|screen|flow|section|area|функція|модуль|сторінка|екран|розділ)\b",
    re.IGNORECASE,
)

# Words that MUST NOT become module names (table noise, summaries, generic labels)
_BLACKLIST = {
    # English
    "total", "subtotal", "summary", "sum", "count", "overall", "grand total",
    "n/a", "na", "none", "tbd", "tba", "yes", "no", "ok", "cancel",
    "module", "page", "feature", "section", "comment", "comments", "description",
    "test cases", "test-cases", "hours", "rate", "version", "author", "date",
    "project", "notes", "note", "item", "items", "row", "column",
    # Ukrainian
    "разом", "підсумок", "усього", "всього", "назва", "опис", "коментар",
    "коментарі", "сторінка", "модуль", "версія", "дата", "автор", "проєкт",
    # Months / weekdays (table noise)
    "january", "february", "march", "april", "may", "june", "july",
    "august", "september", "october", "november", "december",
    "monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday",
    "січень", "лютий", "березень", "квітень", "травень", "червень", "липень",
    "серпень", "вересень", "жовтень", "листопад", "грудень",
}

# Purely generic single words that could be modules ONLY if paired with another word
_GENERIC_SINGLES = {
    "user", "users", "admin", "home", "main", "system", "data", "info",
    "details", "general", "new", "old", "list", "view", "edit", "delete",
    "add", "create", "update", "remove", "save", "submit", "search",
}


def _estimate_tc_for_feature(name: str) -> int:
    """Heuristic: estimate number of test cases for a feature based on its label.

    Uses weights similar to the reference Features sheet (most items 8-10,
    heavier ones like Settings/Payments 15, simple Auth 6).
    """
    low = name.lower()
    if any(k in low for k in ("payment", "checkout", "billing", "subscription",
                               "settings", "admin", "dashboard", "reporting",
                               "analytics", "order", "product", "financial")):
        return 15
    if any(k in low for k in ("auth", "login", "registration", "sign up", "sign in",
                               "authorization", "password")):
        return 6
    if any(k in low for k in ("chat", "messaging", "comment", "notification",
                               "integration", "workspace", "template",
                               "banner", "advertising", "menu", "offers",
                               "delivery")):
        return 10
    if any(k in low for k in ("search", "filter", "sort", "user stories", "stories")):
        return 8
    if any(k in low for k in ("folder", "share", "sharing", "upload", "download",
                               "profile", "calendar", "redesign")):
        return 7
    return 8  # default


def _normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip(" -–—:;.•●|")).strip()


def _is_blacklisted(name: str) -> bool:
    low = name.lower().strip()
    if low in _BLACKLIST:
        return True
    # Pure digits, version strings, bare bullets
    if re.fullmatch(r"[\d\s.,%$/+-]+", low):
        return True
    if re.fullmatch(r"v\d+(\.\d+)*", low):
        return True
    if len(low) < 2:
        return True
    return False


def _looks_like_module(name: str) -> bool:
    """Return True if the string plausibly represents a module / page / feature."""
    name = name.strip()
    if not name or _is_blacklisted(name):
        return False
    if len(name) > 120:
        return False

    words = name.split()
    # Multi-word titles — accept if not too long
    if len(words) >= 2:
        return True
    # Single word — accept only if it is NOT a generic noun
    if name.lower() in _GENERIC_SINGLES:
        return False
    # Accept a single word if it contains a slash (e.g. "Banners/Advertising")
    # or is long enough (5+ chars) and not in the blacklist.
    if "/" in name or len(name) >= 5:
        return True
    return False


def _merge_split_title_lines(lines: list[str]) -> list[str]:
    """PDFs/Docs often break a title like 'User Stories' across two lines.

    Collapse consecutive short Title-Case single-word lines that do not look like
    stand-alone modules into one line.
    """
    merged: list[str] = []
    buf: list[str] = []

    def flush():
        if buf:
            merged.append(" ".join(buf))
            buf.clear()

    for ln in lines:
        s = ln.strip()
        if not s:
            flush(); continue
        # Bulleted or tabular → never merge
        if _ITEM_RE.match(s) or _TABLE_RE.match(s):
            flush(); merged.append(s); continue

        words = s.split()
        if (len(words) == 1 and s[0].isalpha() and s[0].isupper()
                and len(s) <= 20 and s.lower() not in _BLACKLIST):
            buf.append(s)
            if len(buf) >= 4:
                flush()
        else:
            flush(); merged.append(s)
    flush()
    return merged


def features_from_text(text: str) -> list[Feature]:
    """Parse free-form or semi-tabular text and derive features.

    Priority order:
      1. Tabular lines ``Name | NN | comment`` (direct TC count from xlsx/csv).
      2. Bulleted / numbered items under an optional section header.
      3. Plausible module-name lines (multi-word or feature-hinted).

    Every returned Feature has ``test_cases > 0`` unless it is explicitly a section.
    Duplicates (case-insensitive) are removed.
    """
    raw_lines = [ln for ln in text.splitlines()]
    lines = _merge_split_title_lines(raw_lines)

    features: list[Feature] = []
    seen: set[str] = set()
    current_section: str | None = None

    def _add(name: str, tc: int, comment: str = "", is_section: bool = False):
        nm = _normalize_name(name)
        if not nm or _is_blacklisted(nm):
            return
        if not is_section and not _looks_like_module(nm):
            return
        key = (current_section or "") + "::" + nm.lower()
        if key in seen:
            return
        seen.add(key)
        display = f"{current_section} — {nm}" if (current_section and not is_section) else nm
        features.append(Feature(
            name=display,
            test_cases=0 if is_section else max(int(tc), 1),
            comment=comment,
            is_section=is_section,
        ))

    for line in lines:
        s = line.strip()
        if not s:
            continue

        # 1. Tabular row: "Name | 55 | comment"
        tm = _TABLE_RE.match(s)
        if tm:
            name = tm.group(1).strip()
            # Skip header row
            if name.lower() in ("module\\page", "module/page", "module",
                                "task name", "feature"):
                continue
            if _is_blacklisted(name):
                continue
            try:
                tc = int(tm.group(2))
            except ValueError:
                tc = _estimate_tc_for_feature(name)
            tc = max(tc, 1)
            comment = (tm.group(3) or "").strip()
            _add(name, tc, comment=comment)
            continue

        # 2. Bulleted / numbered item
        bm = _ITEM_RE.match(s)
        if bm:
            name = _normalize_name(bm.group(1))
            if name and not _is_blacklisted(name):
                _add(name, _estimate_tc_for_feature(name))
            continue

        # 3. Possible section header (short, no terminal punctuation)
        if (len(s) <= 60 and not s.endswith((".", "!", "?", ";"))
                and (_FEATURE_HINTS.search(s) or s.isupper()
                     or (s.istitle() and len(s.split()) >= 2))):
            # Section header only if it also isn't a plain module candidate.
            current_section = _normalize_name(s.rstrip(":"))
            _add(current_section, 0, is_section=True)
            continue

        # 4. Otherwise treat as a standalone module line if plausible
        cleaned = _normalize_name(s)
        if _looks_like_module(cleaned):
            _add(cleaned, _estimate_tc_for_feature(cleaned))

    # Drop orphan sections that have no children between them and the next one
    pruned: list[Feature] = []
    for i, f in enumerate(features):
        if f.is_section:
            # Keep a section only if at least one following (before next section)
            # child feature belongs to it
            has_child = False
            for g in features[i + 1:]:
                if g.is_section:
                    break
                has_child = True
                break
            if not has_child:
                continue
        pruned.append(f)

    return pruned


#: How many grids on one page are priced individually. A page rendering
#: eight tables is a report, not eight list surfaces, and charging full
#: list-surface coverage for each would swamp the estimate.
MAX_PRICED_GRIDS_PER_PAGE = 3

#: Ceiling on one page's grid budget. Generous — a rich admin grid
#: prices at ~20 — but it bounds a reference/documentation site whose
#: every page ships several tables. Logged when it bites.
MAX_GRID_TC_PER_PAGE = 60


def _grid_tc(page) -> tuple[int, str]:
    """Test-case budget for the grids on one page, and a note naming it.

    Priced by :func:`engine.tc_rules.count_grid_cases`, which walks the
    same ``coverage_rules.yaml`` checks the generator writes from. A
    density formula here would be a second copy of the coverage model,
    and the estimate would drift from the pack the first time a check
    was added to the YAML.

    Returns ``(0, "")`` when the page has no grid, when the crawl
    predates grid support, or when the rules asset is unusable — an
    estimate that silently loses its grid budget is worse than one that
    never had it, so the failure is logged.
    """
    tables = list(getattr(page, "tables", None) or [])
    if not tables:
        return 0, ""
    controls = getattr(page, "grid_controls", None) or {}

    try:
        from engine.tc_rules import count_grid_cases
    except Exception as exc:  # pragma: no cover — defensive
        _logger.warning("qa_estimator: cannot price grids: %s", exc)
        return 0, ""

    priced = tables[:MAX_PRICED_GRIDS_PER_PAGE]
    if len(tables) > len(priced):
        _logger.info("qa_estimator: page has %d grids, priced %d (cap %d)",
                     len(tables), len(priced), MAX_PRICED_GRIDS_PER_PAGE)
    total = sum(count_grid_cases(t, controls) for t in priced)
    if total > MAX_GRID_TC_PER_PAGE:
        _logger.info("qa_estimator: page grid budget %d, kept %d (cap %d)",
                     total, MAX_GRID_TC_PER_PAGE, MAX_GRID_TC_PER_PAGE)
        total = MAX_GRID_TC_PER_PAGE
    if not total:
        return 0, ""
    noun = "grid" if len(priced) == 1 else "grids"
    return total, f"{len(priced)} {noun} (+{total} list-surface cases)"


def features_from_site_analysis(analysis) -> list[Feature]:
    """Build features from a `SiteAnalysis` produced by engine.site_crawler.

    Test-case budget is now *architecture-aware* so a marketing WordPress
    site doesn't get the same per-page count as an e-commerce catalogue
    or a dashboard SaaS. The QA Team Lead reasons like this:

    * WordPress / static / landing — header, navigation and footer are shared
      across most pages, so those get tested once at the global level; each
      content page gets just a small set of content/SEO/responsiveness TCs.
    * SPA — pages are real screens with client-side state, so interactive
      density (forms/buttons/nav) drives the count, but capped per page.
    * E-commerce — product listing, cart and checkout get explicit budgets.
    * Dashboard / admin SaaS — each page is a data view with filters and
      actions; density counts but with a reasonable ceiling.

    Reference reading (bundled in Estimation materials_updated.docx):
      • Three-point / PERT estimation (wideband Delphi, beta distribution)
      • DOU articles on QA estimation accuracy
      • Brooks's law — why adding more testers late doesn't rescue a project

    The estimator still exposes MIN/MAX/Expected via the task pipeline;
    this function only decides how many *test cases* a feature is worth.
    """
    site_type = getattr(analysis, "site_type", "generic") or "generic"
    features: list[Feature] = []
    seen: set[str] = set()

    # --- 1) Global features (shared across pages) ----------------------
    # For shared-template sites these get the bulk of the TC budget so we
    # don't double-count header/footer/nav on every content page.
    global_budget = {
        "wordpress": {"web_general": 8, "auth": 10, "search": 6, "forms": 6, "grids": 5, "payment": 12},
        "static":    {"web_general": 6, "auth": 8,  "search": 5, "forms": 5, "grids": 4, "payment": 10},
        "landing":   {"web_general": 6, "auth": 8,  "search": 5, "forms": 6, "grids": 4, "payment": 10},
        "spa":       {"web_general": 10, "auth": 14, "search": 8, "forms": 8, "grids": 8, "payment": 14},
        "ecommerce": {"web_general": 12, "auth": 16, "search": 10, "forms": 10, "grids": 8, "payment": 22},
        "dashboard": {"web_general": 10, "auth": 16, "search": 8, "forms": 10, "grids": 10, "payment": 14},
        "app":       {"web_general": 10, "auth": 14, "search": 8, "forms": 10, "grids": 8, "payment": 14},
        "generic":   {"web_general": 8, "auth": 10, "search": 6, "forms": 6, "grids": 6, "payment": 12},
    }.get(site_type, None) or {"web_general": 8}

    if getattr(analysis, "features_detected", None):
        for flag in analysis.features_detected:
            key = f"*global*:{flag}".lower()
            if key in seen:
                continue
            seen.add(key)
            tc = global_budget.get(flag, 6)
            features.append(Feature(
                name=f"Global — {flag}",
                test_cases=tc,
                comment=f"shared across site (type={site_type})",
            ))

    # --- 2) Per-page features with architecture-aware density ----------
    # Density formula differs per type; content-heavy sites get a small
    # flat budget, interactive sites scale with form/button density.
    def _per_page_tc(page, interactive: bool) -> tuple[int, str]:
        forms = len(page.forms or [])
        buttons = len(page.buttons or [])
        nav = len(page.nav_links or [])
        grid_tc, grid_note = _grid_tc(page)
        if site_type in ("wordpress", "static", "landing"):
            # Content pages: SEO + copy + responsive + links. Forms add a bit.
            tc = 3 + min(forms, 2) * 2
            cap = 7
            comment = f"content page (forms={forms})"
        elif site_type == "ecommerce":
            tc = 5 + forms * 3 + min(buttons, 10) * 1 + min(nav, 10) * 0.4
            cap = 22
            comment = f"ecommerce page (forms={forms}, buttons={buttons})"
        elif site_type == "dashboard":
            tc = 6 + forms * 2 + min(buttons, 12) * 0.8
            cap = 20
            comment = f"dashboard view (forms={forms}, buttons={buttons})"
        elif site_type == "spa":
            tc = 4 + forms * 3 + min(buttons, 10) * 0.6 + min(nav, 8) * 0.3
            cap = 15
            comment = f"SPA screen (forms={forms}, buttons={buttons})"
        else:  # app / generic
            tc = 4 + forms * 3 + min(buttons, 8) * 0.6 + min(nav, 8) * 0.3
            cap = 14
            comment = f"page (forms={forms}, buttons={buttons})"
        tc = int(round(min(tc, cap)))
        tc = max(tc, 2)
        # The grid budget sits OUTSIDE the density cap on purpose. The
        # cap exists to stop form/button counts running away; a list
        # surface is a separate surface worth 12-20 cases in the
        # reference corpus, and folding it under the same ceiling is
        # exactly how a grid-heavy admin app came out under-estimated.
        if grid_tc:
            tc += grid_tc
            comment = f"{comment}; {grid_note}"
        return tc, comment

    interactive_types = {"spa", "ecommerce", "dashboard", "app"}
    interactive = site_type in interactive_types

    # Per-page cap = how many unique pages we surface as Features rows.
    # Operator-reported on 2026-05-04: testfort.com got only 10 pages
    # in the result table — too shallow even for a low-level checklist.
    # Was: WordPress 8 / static 6 / landing 3 / SPA 12 / app 12 /
    #       dashboard 15 / ecommerce 18 / generic 10.
    # Now bumped roughly 2-3x because a real estimation table for a
    # marketing or e-commerce site needs to enumerate every page that
    # could harbour testable behaviour. The crawler's MAX_PAGES (50)
    # is the upper ceiling — these caps are just the per-architecture
    # de-duplication threshold.
    per_page_cap_count = {
        "wordpress": 25, "static": 20, "landing": 8,
        "spa": 30, "app": 30, "dashboard": 35,
        "ecommerce": 40, "generic": 25,
    }.get(site_type, 25)

    unique_pages: list[tuple] = []
    for page in getattr(analysis, "pages", []) or []:
        title = (page.h1 or page.title or page.url or "Page").strip()
        key = title.lower()
        if key in seen:
            continue
        seen.add(key)
        unique_pages.append((title, page))

    unique_pages = unique_pages[:per_page_cap_count]
    for title, page in unique_pages:
        tc, note = _per_page_tc(page, interactive)
        features.append(Feature(name=title, test_cases=tc, comment=note))

    # --- 3) Architecture summary note ---------------------------------
    notes = getattr(analysis, "architecture_notes", None) or []
    notes_text = "; ".join(notes)
    if notes_text:
        features.insert(0, Feature(
            name=f"Architecture: {site_type}",
            test_cases=0,
            is_section=True,
            comment=notes_text,
        ))

    return features


# ── XLSX export ──────────────────────────────────────────────────

def export_estimation_xlsx(result: EstimationResult, output_path: str) -> str:
    """Export an EstimationResult to a workbook matching the reference layout."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # ==== Sheet 1: Manual QA ====
    ws = wb.active
    ws.title = "Manual QA"

    bold = Font(name="Arial", bold=True)
    base_font = Font(name="Arial")
    head_fill = PatternFill("solid", start_color="E6E6E6")
    highlight = PatternFill("solid", start_color="FFF2CC")
    thin = Side(border_style="thin", color="999999")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header block
    ws["C3"] = "Project:"
    ws["D3"] = result.project_name or "Project Name"
    ws["C4"] = "Version:"
    ws["D4"] = "v1.0"
    ws["C5"] = "Author:"
    ws["D5"] = "TestForTge — QA Team Lead"
    ws["C6"] = "Date:"
    ws["D6"] = result.created_at

    ws["C8"] = "Estimation for Manual QA/Testing"
    ws["C8"].font = Font(name="Arial", bold=True, size=14)

    headers = ["#", "", "Task Name", "Description", "MIN Hours", "MAX Hours",
               "Expected Hours", "Rate USD/Hour"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=10, column=col, value=h)
        c.font = bold
        c.fill = head_fill
        c.border = box
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Rows 11..18
    ROW_KEYS = ["communication", "documentation", "checklist", "functional",
                "bug_report", "bug_recheck", "regression", "compatibility"]
    task_by_key = {t.key: t for t in result.tasks}
    row = 11
    for i, key in enumerate(ROW_KEYS, start=1):
        task = task_by_key[key]
        if key in ("documentation", "functional"):
            ws.cell(row=row, column=2, value=i - 1)
        ws.cell(row=row, column=3, value=task.title)
        ws.cell(row=row, column=4, value=task.description)
        ws.cell(row=row, column=5, value=round(task.min_h, 2))
        ws.cell(row=row, column=6, value=round(task.max_h, 2))
        ws.cell(row=row, column=7, value=f"=AVERAGE(E{row}:F{row})")
        if key == "regression":
            ws.cell(row=row, column=7, value=f"=G{row-3}")  # G14
        for col in range(2, 9):
            ws.cell(row=row, column=col).border = box
            ws.cell(row=row, column=col).font = base_font
        ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    # Row 19: Testing/QA (One Platform)
    r19 = 19
    ws.cell(row=r19, column=3, value="Testing/QA (One Platform)").font = bold
    ws.cell(row=r19, column=5, value="=SUM(E11:E17)")
    ws.cell(row=r19, column=6, value="=SUM(F11:F17)")
    ws.cell(row=r19, column=7, value="=SUM(G11:G17)")

    # Row 20: PM
    r20 = 20
    ws.cell(row=r20, column=3, value="Project management")
    ws.cell(row=r20, column=5, value=f"=E{r19}*{result.pm_overhead}")
    ws.cell(row=r20, column=6, value=f"=F{r19}*{result.pm_overhead}")
    ws.cell(row=r20, column=7, value=f"=G{r19}*{result.pm_overhead}")

    # Row 21: Total
    r21 = 21
    ws.cell(row=r21, column=3, value="Total").font = bold
    ws.cell(row=r21, column=5, value=f"=SUM(E{r19}:E{r20})")
    ws.cell(row=r21, column=6, value=f"=SUM(F{r19}:F{r20})")
    ws.cell(row=r21, column=7, value=f"=SUM(G{r19}:G{r20})")
    for col in (5, 6, 7):
        ws.cell(row=r21, column=col).fill = highlight

    # Row 23: Testing/QA (Full Compatibility)
    r23 = 23
    ws.cell(row=r23, column=2, value="    Testing/QA (Full  compatibility)").font = bold
    ws.cell(row=r23, column=5, value=f"=SUM(E11:E17)+E18*$A$34")
    ws.cell(row=r23, column=6, value=f"=SUM(F11:F17)+F18*$A$34")
    ws.cell(row=r23, column=7, value=f"=AVERAGE(E{r23},F{r23})")

    # Row 24: PM
    r24 = 24
    ws.cell(row=r24, column=3, value="Project management")
    ws.cell(row=r24, column=5, value=f"=E{r23}*{result.pm_overhead}")
    ws.cell(row=r24, column=6, value=f"=F{r23}*{result.pm_overhead}")
    ws.cell(row=r24, column=7, value=f"=G{r23}*{result.pm_overhead}")

    # Row 25: Total
    r25 = 25
    ws.cell(row=r25, column=3, value="Total").font = bold
    ws.cell(row=r25, column=5, value=f"=SUM(E{r23},E{r24})")
    ws.cell(row=r25, column=6, value=f"=SUM(F{r23},F{r24})")
    ws.cell(row=r25, column=7, value=f"=SUM(G{r23},G{r24})")
    for col in (5, 6, 7):
        ws.cell(row=r25, column=col).fill = highlight

    # Cost block
    r27 = 27
    ws.cell(row=r27, column=3, value="Total cost, USD").font = bold
    ws.cell(row=r27, column=5, value="MIN").font = bold
    ws.cell(row=r27, column=6, value="MAX").font = bold
    ws.cell(row=r27, column=7, value="Expected").font = bold
    r28 = 28
    ws.cell(row=r28, column=3, value="Cost Without Compatibility")
    ws.cell(row=r28, column=5, value=f"=E{r21}*$H$11")
    ws.cell(row=r28, column=6, value=f"=F{r21}*$H$11")
    ws.cell(row=r28, column=7, value=f"=G{r21}*$H$11")
    r29 = 29
    ws.cell(row=r29, column=3, value="Cost With Compatibility")
    ws.cell(row=r29, column=5, value=f"=E{r25}*$H$11")
    ws.cell(row=r29, column=6, value=f"=F{r25}*$H$11")
    ws.cell(row=r29, column=7, value=f"=G{r25}*$H$11")

    # Rate in H11 (used by cost formulas)
    ws["H11"] = result.rate_usd or 0

    # A34 — number of additional platforms
    ws["A34"] = result.additional_platforms
    ws["C34"] = (f"** Compatibility testing will be performed on "
                  f"{result.additional_platforms} additional combinations:\n- "
                  + "\n- ".join(result.platforms_list) if result.platforms_list else
                  f"** Compatibility testing will be performed on "
                  f"{result.additional_platforms} additional combinations.")
    ws["C33"] = (f"* Functional Testing & UX Testing will be performed on the "
                  f"following most popular combination: {result.primary_platform}.")
    ws["C32"] = "Notes:"
    ws["C32"].font = bold

    # Column widths
    widths = {"A": 6, "B": 6, "C": 38, "D": 55, "E": 12, "F": 12, "G": 16, "H": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ==== Sheet 2: Features ====
    ws2 = wb.create_sheet("Features")
    ws2["A3"] = "Module\\Page"
    ws2["B3"] = "Test-cases"
    ws2["C3"] = "Comments"
    for c in ("A3", "B3", "C3"):
        ws2[c].font = bold
        ws2[c].fill = head_fill
        ws2[c].border = box

    r = 4
    first_feature_row = r
    for f in result.features:
        if f.is_section:
            cell = ws2.cell(row=r, column=1, value=f.name)
            cell.font = Font(name="Arial", bold=True)
        else:
            ws2.cell(row=r, column=1, value=f.name).font = base_font
            ws2.cell(row=r, column=2, value=int(f.test_cases)).font = base_font
            if f.comment:
                ws2.cell(row=r, column=3, value=f.comment).font = base_font
        r += 1
    last_feature_row = r - 1

    # Totals
    r_total = r
    ws2.cell(row=r_total, column=1, value="Total").font = bold
    ws2.cell(row=r_total, column=2,
             value=f"=SUM(B{first_feature_row}:B{last_feature_row})").font = bold

    r_hours = r_total + 1
    ws2.cell(row=r_hours, column=1, value="Total (hours)").font = bold
    # Hours formula — plain arithmetic (no custom Python int()); matches B40 output
    ws2.cell(
        row=r_hours, column=2,
        value=f"=INT({result.buffer}*B{r_total}*{result.minutes_per_tc}/60)+1",
    ).font = bold
    ws2.cell(row=r_hours, column=3,
             value=f"Estimating {result.minutes_per_tc} minutes per a test case on average, "
                   f"buffer {int(round((result.buffer-1)*100))}%")

    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 60

    # Cross-link from Manual QA E13 to Features!B(r_hours)  — mirror the reference
    ws["E13"] = f"=Features!B{r_hours}"
    ws["F13"] = f"={result.max_testing_stretch}*E13"
    ws["E14"] = "=E13"
    ws["F14"] = "=F13"
    ws["E15"] = f"=(E14+E17)*{result.bug_report_rate}"
    ws["F15"] = f"=(F14+F17)*{result.bug_report_rate}"
    ws["E16"] = f"=(E14+E17)*{result.bug_report_rate}"
    ws["F16"] = f"=(F14+F17)*{result.bug_report_rate}"
    ws["E17"] = "=E14"
    ws["F17"] = "=F14"
    ws["G17"] = "=G14"
    ws["E18"] = f"=E13*{result.compatibility_rate}"
    ws["F18"] = f"=F13*{result.compatibility_rate}"

    wb.save(output_path)
    return output_path
